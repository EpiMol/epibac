#!/usr/bin/env python3
"""
Script para generar el reporte GESTLAB con formato específico y salida en Excel y CSV.

Autor: GitHub Copilot
Fecha: 6 de junio de 2025
"""

import pandas as pd
import os
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def setup_logging():
    """Configurar logging básico."""
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)


def get_species_conversion_dict():
    """Convierte nombres de carpetas abreviados a nombres de especie completos."""
    return {
        'abaumannii': 'Acinetobacter baumannii',
        'citrobacter': 'Citrobacter sp.',
        'cjejuni': 'Campylobacter jejuni',
        'ecoli': 'Escherichia coli',
        'ecoli_achtman': 'Escherichia coli',
        'koxytoca': 'Klebsiella oxytoca',
        'kpneumoniae': 'Klebsiella pneumoniae',
        'lpneumophila': 'Legionella pneumophila',
        'neisseria': 'Neisseria sp.',
        'pputida': 'Pseudomonas putida',
        'saureus': 'Staphylococcus aureus',
        'senterica': 'Salmonella enterica',
        'smaltophilia': 'Stenotrophomonas maltophilia',
        'ecloacae': 'Enterobacter cloacae',
        'efaecium': 'Enterococcus faecium',
        'lmonocytogenes': 'Listeria monocytogenes',
        'mabscessus': 'Mycobacterium abscessus',
        'mycobacteria': 'Mycobacterium sp.',
        'paeruginosa': 'Pseudomonas aeruginosa',
        'proteus': 'Proteus sp.',
        'serratia': 'Serratia sp.'
    }


def convert_species_name(species_name):
    """Convertir nombre de especie abreviado a nombre completo."""
    if pd.isna(species_name) or not species_name or str(species_name).strip() == "" or str(species_name).strip() == "-":
        return "ND"
    
    species_dict = get_species_conversion_dict()
    species_lower = str(species_name).lower().strip()
    
    # Casos especiales
    if species_lower == "-" or species_lower == "":
        return "ND"
    
    # Buscar coincidencia exacta
    if species_lower in species_dict:
        return species_dict[species_lower]
    
    # Buscar coincidencia parcial
    for key, value in species_dict.items():
        if key in species_lower or species_lower in key:
            return value
    
    # Si no hay coincidencia, devolver "ND"
    return "ND"


def generate_plasmid_summary(sample_id, outdir, logger):
    """
    Generar resumen detallado de plásmidos leyendo el archivo MOB-suite individual.

    Formato de salida:
    plasmid_AA002: 87.6kb | conjugative | IncL/M | MOBP+MPF_I | K.pneumoniae(99.0%ANI) | AMRs: none
    plasmid_AA038: 103.6kb | conjugative | IncFII+rep_2183 | MOBF+MPF_F | K.pneumoniae(96.7%ANI) | AMRs: blaTEM-1, aac(6')-Ib
    """
    try:
        logger.info(f"=== Procesando plásmidos para muestra: {sample_id} ===")

        # Construir ruta al archivo MOB-suite individual
        mob_file = os.path.join(outdir, "mge_analysis/plasmids/mob_suite", str(sample_id), "mobtyper_results.txt")
        logger.info(f"Buscando archivo MOB-suite: {mob_file}")

        # Verificar que el archivo existe
        if not os.path.exists(mob_file):
            logger.warning(f"Archivo MOB-suite no encontrado: {mob_file}")
            return "ND"

        # Verificar que el archivo no está vacío
        if os.path.getsize(mob_file) == 0:
            logger.warning(f"Archivo MOB-suite vacío: {mob_file}")
            return "ND"

        # Leer archivo MOB-suite (formato TSV con encabezado duplicado cada 2 filas)
        # Estructura: header (línea 0), data (línea 1), header (línea 2), data (línea 3), ...
        with open(mob_file, 'r') as f:
            lines = f.readlines()

        if len(lines) < 2:
            logger.warning(f"Archivo MOB-suite con datos insuficientes: {mob_file}")
            return "ND"

        # Extraer header de la primera línea
        header = lines[0].strip().split('\t')

        # Extraer solo las filas de datos (líneas 1, 3, 5, 7, ...)
        data_rows = []
        for i in range(1, len(lines), 2):
            data_rows.append(lines[i].strip().split('\t'))

        # Crear DataFrame
        df_plasmids = pd.DataFrame(data_rows, columns=header)
        logger.info(f"Archivo MOB-suite procesado: {len(lines)} líneas, {len(df_plasmids)} plásmidos")

        logger.info(f"Plásmidos encontrados: {len(df_plasmids)}")

        if len(df_plasmids) == 0:
            logger.info(f"Muestra '{sample_id}' no tiene plásmidos predichos")
            return "ND"

        # Generar resumen para cada plásmido
        plasmid_summaries = []

        for idx, row in df_plasmids.iterrows():
            plasmid_id = str(row['sample_id'])
            size_bp = int(row['size'])
            size_kb = size_bp / 1000.0

            # Movilidad
            mobility = str(row['predicted_mobility']).strip()

            # Replicon types (Inc types y otros)
            rep_types = []
            if pd.notna(row['rep_type(s)']) and str(row['rep_type(s)']).strip() not in ['-', '']:
                reps = str(row['rep_type(s)']).split(',')
                rep_types = [r.strip() for r in reps if r.strip()]

            # Relaxase y MPF types
            mob_types = []
            if pd.notna(row['relaxase_type(s)']) and str(row['relaxase_type(s)']).strip() not in ['-', '']:
                relaxases = str(row['relaxase_type(s)']).split(',')
                mob_types.extend([r.strip() for r in relaxases if r.strip()])

            mpf_types = []
            if pd.notna(row['mpf_type']) and str(row['mpf_type']).strip() not in ['-', '']:
                mpfs = str(row['mpf_type']).split(',')
                mpf_types = [m.strip() for m in mpfs if m.strip()]

            # Combinar MOB y MPF (eliminar duplicados preservando orden)
            mob_mpf = []
            seen = set()
            for item in mob_types + mpf_types:
                if item not in seen:
                    mob_mpf.append(item)
                    seen.add(item)
            mob_mpf_str = '+'.join(mob_mpf) if mob_mpf else 'none'

            # Especie y ANI
            species = str(row['mash_neighbor_identification']).strip()
            # Simplificar nombre de especie
            if 'Klebsiella pneumoniae' in species:
                species_short = 'K.pneumoniae'
            elif 'Escherichia coli' in species:
                species_short = 'E.coli'
            elif 'Pseudomonas aeruginosa' in species:
                species_short = 'P.aeruginosa'
            elif 'Acinetobacter baumannii' in species:
                species_short = 'A.baumannii'
            else:
                # Usar primera letra del género + especie
                parts = species.split()
                if len(parts) >= 2:
                    species_short = f"{parts[0][0]}.{parts[1]}"
                else:
                    species_short = species

            # Calcular ANI (1 - mash_distance)
            ani_pct = 0.0
            if pd.notna(row['mash_neighbor_distance']):
                try:
                    mash_dist = float(row['mash_neighbor_distance'])
                    ani_pct = (1 - mash_dist) * 100
                except:
                    ani_pct = 0.0

            species_ani = f"{species_short}({ani_pct:.1f}%ANI)" if ani_pct > 0 else species_short

            # Replicon types como string compacto
            rep_str = '+'.join(rep_types) if rep_types else 'none'

            # Construir resumen del plásmido (sin AMRs, ya que no los estamos evaluando por plásmido)
            summary_parts = [
                plasmid_id,
                f"{size_kb:.1f}kb",
                mobility,
                rep_str,
                mob_mpf_str,
                species_ani
            ]

            plasmid_summary = " | ".join(summary_parts)
            plasmid_summaries.append(plasmid_summary)
            logger.info(f"  {plasmid_summary}")

        # Unir todos los plásmidos con salto de línea
        final_summary = "\n".join(plasmid_summaries)
        logger.info(f"Resumen completo generado para '{sample_id}'")

        return final_summary

    except Exception as e:
        logger.error(f"Error procesando plásmidos para '{sample_id}': {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return "ND"


def process_path_nlsar(file_path):
    """
    Procesa una ruta de archivo para que comience desde NLSAR.
    
    Ejemplo:
    /home/asanzc/CabinaCS/NLSAR/deposit/CVA_EPIM/illumina/250425_EPIM199/fastq/EPI00579_S43_R1_001.fastq.gz
    -> NLSAR/deposit/CVA_EPIM/illumina/250425_EPIM199/fastq/EPI00579_S43_R1_001.fastq.gz
    """
    if pd.isna(file_path) or not file_path:
        return file_path
    
    # Buscar la posición de "NLSAR" en la ruta
    nlsar_pos = file_path.find("NLSAR")
    if nlsar_pos != -1:
        return file_path[nlsar_pos:]
    else:
        # Si no contiene NLSAR, devolver la ruta original
        return file_path


def determine_seq_method(row):
    """Determinar el tipo de secuenciación (ILLUMINA, NANOPORE o HYBRID)."""
    has_illumina = pd.notna(row.get("ILLUMINA_R1", "")) and row.get("ILLUMINA_R1", "") != ""
    has_nanopore = pd.notna(row.get("NANOPORE", "")) and row.get("NANOPORE", "") != ""

    if has_illumina and has_nanopore:
        return "HYBRID"
    elif has_illumina:
        return "ILLUMINA"
    elif has_nanopore:
        return "NANOPORE"
    else:
        return pd.NA


def build_path(row, config, tag_run):
    """Construir la ruta de FICHERO_LECTURAS_WGS."""
    if pd.isna(row.get("CARRERA", "")):
        return pd.NA
    
    parts = row["CARRERA"].split("_")
    if len(parts) < 2:
        return pd.NA
    
    # Obtener los parámetros específicos del modo GVA
    storage_cabinet = (
        config.get("mode_config", {})
        .get("gva", {})
        .get("storage_cabinet", "/mnt/CabinaCS/NLSAR/deposit")
    )
    hosp = parts[1][:4]  # Primeros 4 caracteres del segundo segmento
    
    # Para simplificar, usamos la plataforma primaria para la ruta mostrada
    platform = "illumina" if row["OBS_MET_WGS"] in ["ILLUMINA", "HYBRID"] else "nanopore"
    
    # Construir ruta de destino base - usar forward slash (/) para Linux
    base_path = f"{storage_cabinet}/CVA_{hosp}/{platform}/{row['CARRERA']}/fastq"
    
    # Añadir el nombre del archivo si existe
    if platform == "illumina" and pd.notna(row.get("ILLUMINA_R1", "")) and row.get("ILLUMINA_R1", ""):
        # Usar solo el nombre del archivo, no la ruta completa
        filename = os.path.basename(row["ILLUMINA_R1"])
        full_path = f"{base_path}/{filename}"
    elif platform == "nanopore" and pd.notna(row.get("NANOPORE", "")) and row.get("NANOPORE", ""):
        # Usar solo el nombre del archivo, no la ruta completa
        filename = os.path.basename(row["NANOPORE"])
        full_path = f"{base_path}/{filename}"
    else:
        full_path = base_path
    
    # Procesar la ruta para que comience desde NLSAR
    return process_path_nlsar(full_path)


def clean_dataframe_for_excel(df, logger):
    """Limpiar DataFrame para compatibilidad con Excel."""
    try:
        # Crear una copia del dataframe para no modificar el original
        df_clean = df.copy()
        
        # Convertir valores pandas NA/NaN/None a "ND"
        df_clean = df_clean.fillna('ND')
        
        # Convertir cualquier objeto dtype a string para evitar problemas
        for col in df_clean.columns:
            if df_clean[col].dtype == 'object':
                df_clean[col] = df_clean[col].astype(str)
                # Reemplazar 'nan', '<NA>', 'None' por 'ND'
                df_clean[col] = df_clean[col].replace(['nan', '<NA>', 'None', 'NaN', ''], 'ND')
        
        logger.info("DataFrame limpiado para compatibilidad con Excel - campos vacíos rellenados con 'ND'")
        return df_clean
        
    except Exception as e:
        logger.error(f"Error limpiando DataFrame: {e}")
        return df


def create_excel_with_formatting(df, output_path, logger):
    """Crear archivo Excel con formato específico."""
    try:
        # Limpiar DataFrame antes de escribir a Excel
        df_clean = clean_dataframe_for_excel(df, logger)

        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "GESTLAB_Report"

        # Agregar datos al worksheet
        for r in dataframe_to_rows(df_clean, index=False, header=True):
            ws.append(r)

        # Formatear la primera fila (encabezados)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul clarito

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Ajustar ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Establecer ancho mínimo y máximo
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Aplicar wrap text a todas las celdas
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Archivo Excel guardado exitosamente en: {output_path}")

    except Exception as e:
        logger.error(f"Error creando archivo Excel: {e}")
        raise


def create_plasmids_excel(samples_file, outdir, output_path, logger):
    """
    Crear archivo Excel detallado con información de plásmidos.

    Columnas:
    - Sample_ID
    - Plasmid_ID
    - Size_bp
    - Size_kb
    - Mobility
    - Rep_types
    - Relaxase_types
    - MPF_types
    - Species_match
    - ANI_%
    - Num_contigs
    """
    try:
        logger.info("=== Generando Excel detallado de plásmidos ===")

        # Leer archivo de muestras para obtener lista de IDs
        df_samples = pd.read_csv(samples_file, sep=";")
        # El archivo validated puede tener 'id' o 'CODIGO_MUESTRA_ORIGEN'
        if 'id' in df_samples.columns:
            sample_ids = df_samples['id'].tolist()
        elif 'CODIGO_MUESTRA_ORIGEN' in df_samples.columns:
            sample_ids = df_samples['CODIGO_MUESTRA_ORIGEN'].tolist()
        else:
            raise ValueError("No se encontró columna de ID de muestra ('id' o 'CODIGO_MUESTRA_ORIGEN')")

        all_plasmid_data = []

        for sample_id in sample_ids:
            mob_file = os.path.join(outdir, "mge_analysis/plasmids/mob_suite", str(sample_id), "mobtyper_results.txt")

            if not os.path.exists(mob_file) or os.path.getsize(mob_file) == 0:
                continue

            # Leer MOB-suite file
            with open(mob_file, 'r') as f:
                lines = f.readlines()

            if len(lines) < 2:
                continue

            header = lines[0].strip().split('\t')
            data_rows = []
            for i in range(1, len(lines), 2):
                data_rows.append(lines[i].strip().split('\t'))

            df_plasmids = pd.DataFrame(data_rows, columns=header)

            # Extraer información de cada plásmido
            for idx, row in df_plasmids.iterrows():
                plasmid_id = str(row['sample_id'])
                size_bp = int(row['size'])
                size_kb = round(size_bp / 1000.0, 1)
                mobility = str(row['predicted_mobility']).strip()
                num_contigs = int(row['num_contigs'])

                # Rep types
                rep_types = str(row['rep_type(s)']).strip()
                if rep_types in ['-', '', 'nan']:
                    rep_types = 'none'

                # Relaxase types
                relaxase_types = str(row['relaxase_type(s)']).strip()
                if relaxase_types in ['-', '', 'nan']:
                    relaxase_types = 'none'

                # MPF types
                mpf_types = str(row['mpf_type']).strip()
                if mpf_types in ['-', '', 'nan']:
                    mpf_types = 'none'

                # Species and ANI
                species = str(row['mash_neighbor_identification']).strip()
                ani_pct = 0.0
                if pd.notna(row['mash_neighbor_distance']):
                    try:
                        mash_dist = float(row['mash_neighbor_distance'])
                        ani_pct = round((1 - mash_dist) * 100, 1)
                    except:
                        ani_pct = 0.0

                plasmid_data = {
                    'Sample_ID': sample_id,
                    'Plasmid_ID': plasmid_id,
                    'Size_bp': size_bp,
                    'Size_kb': size_kb,
                    'Mobility': mobility,
                    'Rep_types': rep_types,
                    'Relaxase_types': relaxase_types,
                    'MPF_types': mpf_types,
                    'Species_match': species,
                    'ANI_%': ani_pct,
                    'Num_contigs': num_contigs
                }

                all_plasmid_data.append(plasmid_data)

        if len(all_plasmid_data) == 0:
            logger.warning("No se encontraron plásmidos en ninguna muestra")
            # Crear DataFrame vacío con columnas
            df_plasmids_excel = pd.DataFrame(columns=[
                'Sample_ID', 'Plasmid_ID', 'Size_bp', 'Size_kb', 'Mobility',
                'Rep_types', 'Relaxase_types', 'MPF_types', 'Species_match', 'ANI_%', 'Num_contigs'
            ])
        else:
            df_plasmids_excel = pd.DataFrame(all_plasmid_data)

        # Crear Excel con formato
        wb = Workbook()
        ws = wb.active
        ws.title = "Plasmids_Detail"

        # Agregar datos
        for r in dataframe_to_rows(df_plasmids_excel, index=False, header=True):
            ws.append(r)

        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Ajustar anchos de columna
        column_widths = {
            'A': 20,  # Sample_ID
            'B': 18,  # Plasmid_ID
            'C': 12,  # Size_bp
            'D': 12,  # Size_kb
            'E': 18,  # Mobility
            'F': 25,  # Rep_types
            'G': 20,  # Relaxase_types
            'H': 20,  # MPF_types
            'I': 25,  # Species_match
            'J': 10,  # ANI_%
            'K': 12   # Num_contigs
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Wrap text en todas las celdas
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Guardar
        wb.save(output_path)
        logger.info(f"Excel de plásmidos guardado: {output_path} ({len(all_plasmid_data)} plásmidos)")

    except Exception as e:
        logger.error(f"Error creando Excel de plásmidos: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise


def main():
    """Función principal del script."""
    logger = setup_logging()
    
    # Obtener argumentos del snakemake
    try:
        validated_samples_info = snakemake.input.validated_samples_info
        results_file = snakemake.input.results
        plasmids_file = snakemake.input.plasmids
        output_csv = snakemake.output.gestlab_report_csv
        output_excel = snakemake.output.gestlab_report_xlsx
        tag_run = snakemake.params.tag_run
        outdir = snakemake.params.outdir
        config = snakemake.config
    except NameError:
        logger.error("Este script debe ejecutarse desde Snakemake")
        sys.exit(1)
    
    logger.info("=== Iniciando procesamiento de reporte GESTLAB ===")
    logger.info(f"TAG_RUN: {tag_run}")
    
    try:
        # Leer archivos de entrada
        logger.info("Leyendo archivos de entrada...")
        df_meta = pd.read_csv(validated_samples_info, sep=";")
        df_results = pd.read_csv(results_file, sep="\t")
        
        # Detectar si Sample coincide con id o con id2
        if df_results["Sample"].isin(df_meta["id"]).all():
            merge_col = "id"
        elif df_results["Sample"].isin(df_meta["id2"]).all():
            merge_col = "id2"
        else:
            raise ValueError(
                "No se puede encontrar coincidencia entre 'Sample' y 'id' o 'id2'."
            )
        
        logger.info(f"Usando columna '{merge_col}' para el merge")
        
        # Realizar merge
        df_results = df_results.rename(columns={"Sample": merge_col})
        df_merged = df_meta.merge(df_results, on=merge_col, how="left")
        
        # Eliminar columnas no deseadas
        columns_to_remove = ["SCOPE_core", "GENE_resfinder", "PHENO_resfinder", "MLST"]
        for col in columns_to_remove:
            if col in df_merged.columns:
                df_merged.drop(columns=col, inplace=True)
                logger.info(f"Eliminada columna: {col}")
        
        # Añadir CARRERA
        df_merged["CARRERA"] = tag_run
        
        # Renombrar columnas según el mapa GVA
        rename_map_gva = {
            "CODIGO_MUESTRA_ORIGEN": "id",
            "PETICION": "id2",
            "FECHA_TOMA_MUESTRA": "collection_date",
            "ESPECIE_SECUENCIA": "organism",
            "MOTIVO_WGS": "relevance",
            "ILLUMINA_R1": "illumina_r1",
            "ILLUMINA_R2": "illumina_r2",
            "NANOPORE": "nanopore",
            "ID_WGS": "Scheme_mlst",  # Cambiado de ID_WS a ID_WGS
            "ST_WGS": "ST",
            "MLST_WGS": "MLST",
            "R_Geno_WGS": "AMR",
            "PHENO_WGS": "PHENO_resfinder",
            "V_WGS": "VIRULENCE",
            "CONFIRMACION": "confirmation_note",
            "NUM_BROTE": "outbreak_id",
            "COMENTARIO_WGS": "comment",
        }
        
        df_merged.rename(
            columns={v: k for k, v in rename_map_gva.items()}, inplace=True
        )
        
        # Convertir nombres de especies (aplicar solo a ID_WGS)
        if "ID_WGS" in df_merged.columns:
            df_merged["ID_WGS"] = df_merged["ID_WGS"].apply(convert_species_name)
            logger.info("Nombres de especies convertidos en ID_WGS")
        
        # ESPECIE_SECUENCIA mantiene su valor original del samplesheet
        
        # Generar resumen de plásmidos
        logger.info("Generando resumen de plásmidos...")
        logger.info(f"Columnas disponibles después del merge: {list(df_merged.columns)}")
        
        # Determinar qué columna usar para el ID de muestra
        # En el modo GVA, CODIGO_MUESTRA_ORIGEN se mapea a 'id' después del rename
        # y debe coincidir con Sample_ID en el archivo de plásmidos
        id_column = None
        if "CODIGO_MUESTRA_ORIGEN" in df_merged.columns:
            id_column = "CODIGO_MUESTRA_ORIGEN"
        elif "id" in df_merged.columns:
            id_column = "id"
        elif merge_col in df_merged.columns:
            id_column = merge_col
        
        if id_column:
            logger.info(f"Usando columna '{id_column}' para buscar plásmidos")
            logger.info(f"Valores en {id_column}: {df_merged[id_column].tolist()}")

            df_merged["PLASMIDOS_WGS"] = df_merged[id_column].apply(
                lambda sample_id: generate_plasmid_summary(sample_id, outdir, logger)
            )
            logger.info(f"Resumen de plásmidos generado usando columna '{id_column}'")
        else:
            logger.warning("No se encontró columna de ID para generar resumen de plásmidos")
            df_merged["PLASMIDOS_WGS"] = "ND"
        
        # Añadir columnas extra
        df_merged["MO_EST_WGS"] = "BACTERIA"
        df_merged["TIPO_ANALISIS_WGS"] = "VERIFICACION"
        
        # Determinar el tipo de secuenciación
        df_merged["OBS_MET_WGS"] = df_merged.apply(determine_seq_method, axis=1)
        
        # Generar FICHERO_LECTURAS_WGS
        df_merged["FICHERO_LECTURAS_WGS"] = df_merged.apply(
            lambda row: build_path(row, config, tag_run), axis=1
        )
        
        # Añadir las nuevas columnas requeridas
        df_merged["ENVIAR_REDLABRA"] = "ND"
        df_merged["VERIFICACION_WGS"] = "ND"
        
        # Generar ruta para FICHERO_ANALISIS_WGS
        analysis_file_path = f"epibac/output/{tag_run}/report/{tag_run}_EPIBAC_GESTLAB.csv"
        df_merged["FICHERO_ANALISIS_WGS"] = analysis_file_path
        
        # Eliminar las columnas especificadas del archivo final
        columns_to_drop_final = ["ILLUMINA_R1", "ILLUMINA_R2", "NANOPORE", "dorado_model"]
        for col in columns_to_drop_final:
            if col in df_merged.columns:
                df_merged.drop(columns=col, inplace=True)
                logger.info(f"Eliminada columna del resultado final: {col}")
        
        # Rellenar campos vacíos con "ND" para ambos archivos de salida
        df_merged = df_merged.fillna('ND')
        for col in df_merged.columns:
            if df_merged[col].dtype == 'object':
                df_merged[col] = df_merged[col].astype(str)
                df_merged[col] = df_merged[col].replace(['nan', '<NA>', 'None', 'NaN', ''], 'ND')
        
        logger.info("Campos vacíos rellenados con 'ND' en el DataFrame final")
        
        # Guardar archivo CSV
        df_merged.to_csv(output_csv, sep=";", index=False)
        logger.info(f"Archivo CSV guardado en: {output_csv}")
        
        # Crear archivo Excel con formato
        create_excel_with_formatting(df_merged, output_excel, logger)

        # Crear Excel detallado de plásmidos
        plasmids_excel_path = output_excel.replace("_GESTLAB.xlsx", "_Plasmids_Detail.xlsx")
        create_plasmids_excel(validated_samples_info, outdir, plasmids_excel_path, logger)

        # Estadísticas
        logger.info("=== Estadísticas finales ===")
        logger.info(f"Total de muestras procesadas: {len(df_merged)}")
        logger.info(f"Columnas en el archivo final: {len(df_merged.columns)}")
        logger.info(f"Columnas: {list(df_merged.columns)}")

        logger.info("=== Procesamiento completado exitosamente ===")

    except Exception as e:
        logger.error(f"Error durante el procesamiento: {e}")
        raise


if __name__ == "__main__":
    main()
